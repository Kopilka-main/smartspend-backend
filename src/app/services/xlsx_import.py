import json
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.models.card import Card
from src.app.models.company import Company
from src.app.models.deposit import Deposit
from src.app.models.promo import Promo


class FieldType(str, Enum):
    STR = "str"
    INT = "int"
    BOOL = "bool"
    JSON = "json"
    STR_ARRAY = "str_array"
    DATETIME = "datetime"


@dataclass(frozen=True)
class Col:
    column: str
    field: str
    type: FieldType
    required: bool = False


@dataclass(frozen=True)
class EntityConfig:
    model: type
    id_field: str
    columns: list[Col]


@dataclass
class RowError:
    row: int
    message: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list[RowError] = field(default_factory=list)


CARD = EntityConfig(
    model=Card,
    id_field="id",
    columns=[
        Col("A", "id", FieldType.STR, required=True),
        Col("B", "bank_name", FieldType.STR, required=True),
        Col("C", "name", FieldType.STR, required=True),
        Col("D", "card_type", FieldType.STR, required=True),
        Col("E", "cashback", FieldType.JSON),
        Col("F", "cashback_note", FieldType.STR),
        Col("G", "grace_days", FieldType.INT),
        Col("H", "fee", FieldType.INT),
        Col("I", "fee_base", FieldType.INT),
        Col("J", "conditions", FieldType.STR_ARRAY),
        Col("K", "tags", FieldType.STR_ARRAY),
        Col("L", "url", FieldType.STR),
        Col("M", "is_active", FieldType.BOOL),
        Col("N", "bank_logo_url", FieldType.STR),
        Col("O", "bonus_type", FieldType.STR),
        Col("P", "bonus_system", FieldType.STR),
        Col("Q", "bonus_desc", FieldType.STR),
    ],
)

DEPOSIT = EntityConfig(
    model=Deposit,
    id_field="id",
    columns=[
        Col("A", "id", FieldType.STR, required=True),
        Col("B", "bank_name", FieldType.STR, required=True),
        Col("C", "name", FieldType.STR, required=True),
        Col("D", "rates", FieldType.JSON, required=True),
        Col("E", "min_amount", FieldType.INT),
        Col("F", "max_amount", FieldType.INT),
        Col("G", "freq", FieldType.STR),
        Col("H", "replenishment", FieldType.BOOL),
        Col("I", "withdrawal", FieldType.BOOL),
        Col("J", "conditions", FieldType.STR_ARRAY),
        Col("K", "tags", FieldType.STR_ARRAY),
        Col("L", "is_active", FieldType.BOOL),
        Col("M", "bank_logo_url", FieldType.STR),
        Col("N", "ear", FieldType.JSON),
        Col("O", "income_coef", FieldType.JSON),
    ],
)

COMPANY = EntityConfig(
    model=Company,
    id_field="id",
    columns=[
        Col("A", "id", FieldType.STR, required=True),
        Col("B", "name", FieldType.STR, required=True),
        Col("C", "abbr", FieldType.STR),
        Col("D", "color", FieldType.STR),
        Col("E", "category_id", FieldType.STR),
        Col("F", "description", FieldType.STR),
        Col("G", "promo_types", FieldType.STR_ARRAY),
        Col("H", "is_active", FieldType.BOOL),
        Col("I", "logo_url", FieldType.STR),
    ],
)

PROMO = EntityConfig(
    model=Promo,
    id_field="id",
    columns=[
        Col("A", "id", FieldType.INT, required=True),
        Col("B", "type", FieldType.STR, required=True),
        Col("C", "company_id", FieldType.STR),
        Col("D", "category_id", FieldType.STR),
        Col("E", "author_id", FieldType.STR),
        Col("F", "title", FieldType.STR),
        Col("G", "text", FieldType.STR, required=True),
        Col("H", "code", FieldType.STR),
        Col("I", "url", FieldType.STR),
        Col("J", "source_url", FieldType.STR),
        Col("K", "promo_filter", FieldType.STR),
        Col("L", "conditions", FieldType.STR_ARRAY),
        Col("M", "expires_at", FieldType.DATETIME),
        Col("N", "is_active", FieldType.BOOL),
        Col("O", "partner_company_id", FieldType.STR),
    ],
)

CONFIGS: dict[str, EntityConfig] = {
    "cards": CARD,
    "deposits": DEPOSIT,
    "companies": COMPANY,
    "promos": PROMO,
}


def _parse(value: Any, t: FieldType) -> Any:
    if value is None:
        return None
    if t is FieldType.STR:
        return str(value).strip() or None
    if t is FieldType.INT:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
        return int(value)
    if t is FieldType.BOOL:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().lower()
        if text in ("1", "true", "да", "yes"):
            return True
        if text in ("0", "false", "нет", "no", ""):
            return False
        raise ValueError(f"bad boolean: {value}")
    if t is FieldType.DATETIME:
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        if not text:
            return None
        return datetime.fromisoformat(text)
    if t is FieldType.JSON:
        if isinstance(value, (dict, list)):
            return value
        text = str(value).strip()
        if not text:
            return None
        return json.loads(text)
    if t is FieldType.STR_ARRAY:
        if isinstance(value, list):
            return value
        text = str(value).strip()
        if not text:
            return None
        if text.startswith("["):
            parsed = json.loads(text)
            if not isinstance(parsed, list):
                raise ValueError("expected array")
            return parsed
        return [x.strip() for x in text.split(",") if x.strip()]
    return None


async def import_xlsx(session: AsyncSession, entity: str, data: bytes) -> ImportResult:
    config = CONFIGS[entity]
    result = ImportResult()

    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        result.errors.append(RowError(0, "Empty workbook"))
        return result

    col_letters = {c.column for c in config.columns}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_map: dict[str, Any] = {cell.column_letter: cell.value for cell in row if cell.column_letter in col_letters}
        if all(v is None for v in row_map.values()):
            continue

        parsed: dict[str, Any] = {}
        row_error: RowError | None = None

        for col in config.columns:
            raw = row_map.get(col.column)
            if col.required and (raw is None or (isinstance(raw, str) and not raw.strip())):
                row_error = RowError(row_idx, f"Поле {col.field} обязательно")
                break
            try:
                parsed[col.field] = _parse(raw, col.type)
            except (ValueError, json.JSONDecodeError) as exc:
                row_error = RowError(row_idx, f"Поле {col.field} (колонка {col.column}): {exc}")
                break

        if row_error:
            result.errors.append(row_error)
            continue

        record_id = parsed.get(config.id_field)
        existing = (
            await session.execute(select(config.model).where(getattr(config.model, config.id_field) == record_id))
        ).scalar_one_or_none()

        if existing is not None:
            for k, v in parsed.items():
                if k == config.id_field or v is None:
                    continue
                setattr(existing, k, v)
            result.updated += 1
        else:
            session.add(config.model(**{k: v for k, v in parsed.items() if v is not None}))
            result.created += 1

    await session.commit()
    wb.close()
    return result
